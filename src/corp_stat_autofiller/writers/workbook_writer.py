from __future__ import annotations

from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries

from corp_stat_autofiller.models import FillPlanItem


def write_workbook(
    template_path: str | Path,
    output_path: str | Path,
    plan: list[FillPlanItem],
    keep_formulas: bool = True,
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    copy2(template_path, output)
    workbook = load_workbook(output, data_only=not keep_formulas)
    for item in plan:
        if item.sheet not in workbook.sheetnames:
            raise ValueError(f"調査票にシートが見つかりません: {item.sheet}")
        worksheet = workbook[item.sheet]
        worksheet[resolve_writable_cell(worksheet, item.cell)] = item.value
    workbook.save(output)
    return output


def resolve_writable_cell(worksheet, cell: str) -> str:
    for merged_range in worksheet.merged_cells.ranges:
        if cell in merged_range:
            min_col, min_row, _, _ = range_boundaries(str(merged_range))
            return f"{get_column_letter(min_col)}{min_row}"
    return cell
